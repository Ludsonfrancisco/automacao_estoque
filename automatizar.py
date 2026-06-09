"""Fase 2 — Automação em lote dos seriais da planilha.

Uso:
    python automatizar.py

Lê devolucao-27-05.xlsx, processa cada serial da coluna A,
grava Status/Detalhe/Timestamp nas colunas B/C/D.
Salva após cada serial. Pula linhas já preenchidas (retomada).
"""
from __future__ import annotations

import logging
import os
import sys
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

import loga_client


PLANILHA = "DEVOLUCAO-08-06.xlsx"
ABA = "Plan1"
COL_SERIAL = 1
COL_STATUS = 2
COL_DETALHE = 3
COL_TIMESTAMP = 4
HEADER_ROW = 1
FIRST_DATA_ROW = 2


def setup_logging() -> Path:
    Path("logs").mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d")
    log_path = Path("logs") / f"execucao-{stamp}.log"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[logging.FileHandler(log_path, encoding="utf-8"), logging.StreamHandler()],
    )
    return log_path


def ensure_headers(ws) -> None:
    """Garante que B1=Status, C1=Detalhe, D1=Timestamp."""
    headers = {COL_STATUS: "Status", COL_DETALHE: "Detalhe", COL_TIMESTAMP: "Timestamp"}
    for col, header in headers.items():
        if ws.cell(HEADER_ROW, col).value != header:
            ws.cell(HEADER_ROW, col).value = header


def iter_pendentes(ws):
    """Gera (row_idx, serial) para linhas com serial e SEM status preenchido."""
    for row_idx in range(FIRST_DATA_ROW, ws.max_row + 1):
        serial = ws.cell(row_idx, COL_SERIAL).value
        status = ws.cell(row_idx, COL_STATUS).value
        if serial and not status:
            yield row_idx, str(serial).strip()


def grava_resultado(ws, row_idx: int, status: str, detalhe: str | None) -> None:
    ws.cell(row_idx, COL_STATUS).value = status
    ws.cell(row_idx, COL_DETALHE).value = detalhe or ""
    ws.cell(row_idx, COL_TIMESTAMP).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def relogar(page, user: str, password: str) -> bool:
    try:
        loga_client.login(page, user, password)
        loga_client.goto_retorno_materiais(page)
        return True
    except Exception as e:
        logging.error(f"Falha ao relogar: {e}")
        return False


def aguardar_modal_inicial(page) -> None:
    """Aguarda o modal 'Aguarde... Carregando atendimentos...' sumir."""
    try:
        page.locator(".swal2-popup").first.wait_for(state="hidden", timeout=30000)
    except Exception:
        pass


def main() -> int:
    log_path = setup_logging()
    logging.info(f"Log em {log_path}")

    load_dotenv()
    user = os.getenv("LOGA_USER")
    password = os.getenv("LOGA_PASS")
    if not user or not password:
        logging.error("Defina LOGA_USER e LOGA_PASS no .env")
        return 1

    if not Path(PLANILHA).exists():
        logging.error(f"Planilha não encontrada: {PLANILHA}")
        return 1

    wb = load_workbook(PLANILHA)
    ws = wb[ABA]
    ensure_headers(ws)
    wb.save(PLANILHA)

    pendentes = list(iter_pendentes(ws))
    total = len(pendentes)
    logging.info(f"{total} seriais pendentes")
    if total == 0:
        logging.info("Nada a fazer.")
        return 0

    sucessos = 0
    erros = 0
    linhas_com_erro: list[int] = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context()
        page = context.new_page()

        logging.info("Fazendo login...")
        try:
            loga_client.login(page, user, password)
            loga_client.goto_retorno_materiais(page)
            aguardar_modal_inicial(page)
        except Exception as e:
            logging.error(f"Login inicial falhou: {e}")
            browser.close()
            return 2

        logging.info("Pronto. Iniciando processamento...")

        for i, (row_idx, serial) in enumerate(pendentes, 1):
            logging.info(f"[{i}/{total}] linha {row_idx} — serial {serial}")
            try:
                result = loga_client.scan_serial(page, serial)
                grava_resultado(ws, row_idx, result.status, result.detalhe)
                sucessos += 1
                detalhe_log = f" (atend: {result.detalhe})" if result.detalhe else ""
                logging.info(f"  -> {result.status}{detalhe_log}")
            except RuntimeError as e:
                logging.warning(f"  sessão pode ter caído: {e} — tentando relogar")
                if relogar(page, user, password):
                    aguardar_modal_inicial(page)
                    try:
                        result = loga_client.scan_serial(page, serial)
                        grava_resultado(ws, row_idx, result.status, result.detalhe)
                        sucessos += 1
                        logging.info(f"  -> {result.status} (após relogin)")
                    except Exception as e2:
                        msg = f"ERRO: {type(e2).__name__}: {str(e2)[:120]}"
                        grava_resultado(ws, row_idx, msg, None)
                        erros += 1
                        linhas_com_erro.append(row_idx)
                        logging.exception("  falhou após relogin")
                else:
                    grava_resultado(ws, row_idx, "ERRO: relogin falhou", None)
                    erros += 1
                    linhas_com_erro.append(row_idx)
            except Exception as e:
                msg = f"ERRO: {type(e).__name__}: {str(e)[:120]}"
                grava_resultado(ws, row_idx, msg, None)
                erros += 1
                linhas_com_erro.append(row_idx)
                logging.exception(f"  erro processando serial {serial}")

            wb.save(PLANILHA)

        browser.close()

    logging.info("=" * 60)
    logging.info(f"FIM. Total: {total} | Sucessos: {sucessos} | Erros: {erros}")
    if linhas_com_erro:
        logging.info(f"Linhas com erro: {linhas_com_erro}")
    return 0 if erros == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
